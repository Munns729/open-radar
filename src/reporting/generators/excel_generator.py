"""Excel report generator with color-coded cells."""

from typing import List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.universe.database import CompanyModel as Company
from .base import BaseGenerator
from ..filters import get_priority_level


class ExcelReportGenerator(BaseGenerator):
    """Generate Excel reports with color-coded priorities."""
    
    def generate(self, companies: List[Company], filename: str = None) -> str:
        """Generate Excel report."""
        output_file = self.output_dir / self._get_filename("radar_targets", "xlsx", filename)
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Generate sheets
        self._create_summary_sheet(wb, companies)
        self._create_tier_sheet(wb, companies, '1A')
        self._create_tier_sheet(wb, companies, '1B')
        self._create_sector_sheet(wb, companies)
        
        # Save workbook
        wb.save(output_file)
        
        return str(output_file)
    
    def _create_summary_sheet(self, wb: Workbook, companies: List[Company]):
        """Create summary sheet."""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "RADAR Investment Targets Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Statistics
        stats = [
            ("Total Companies", len(companies)),
            ("Tier 1A", sum(1 for c in companies if c.tier == 'TIER_1A')),
            ("Tier 1B", sum(1 for c in companies if c.tier == 'TIER_1B')),
            ("HOT (70+)", sum(1 for c in companies if c.moat_score >= 70)),
            ("HIGH (50-69)", sum(1 for c in companies if 50 <= c.moat_score < 70)),
            ("MED (30-49)", sum(1 for c in companies if 30 <= c.moat_score < 50)),
            ("LOW (<30)", sum(1 for c in companies if c.moat_score < 30)),
            ("Avg Moat Score", f"{sum(c.moat_score for c in companies) / len(companies):.1f}" if companies else "0"),
        ]
        
        for i, (label, value) in enumerate(stats, start=4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _create_tier_sheet(self, wb: Workbook, companies: List[Company], tier: str):
        """Create tier-specific sheet."""
        ws = wb.create_sheet(f"Tier {tier}")
        
        # Headers
        headers = ["Priority", "Company Name", "Moat Score", "Sector", "Revenue", "Moat Type", "Status", "Notes"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Freeze top row
        ws.freeze_panes = "A2"
        
        # Filter companies by tier
        tier_companies = [c for c in companies if c.tier == f"TIER_{tier}"]
        tier_companies.sort(key=lambda x: x.moat_score, reverse=True)
        
        # Color fills for priorities
        priority_fills = {
            'hot': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
            'high': PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid"),
            'med': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            'low': PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
        }
        
        # Data rows
        for row, company in enumerate(tier_companies, start=2):
            priority = get_priority_level(company.moat_score)
            fill = priority_fills[priority]
            
            # Priority column with color
            priority_cell = ws.cell(row, 1, priority.upper())
            priority_cell.fill = fill
            priority_cell.font = Font(bold=True)
            priority_cell.alignment = Alignment(horizontal='center')
            
            # Other columns
            ws.cell(row, 2, company.name)
            ws.cell(row, 3, company.moat_score).alignment = Alignment(horizontal='center')
            ws.cell(row, 4, company.sector or "N/A")
            ws.cell(row, 5, self._format_revenue(company.revenue_gbp))
            ws.cell(row, 6, company.moat_type or "N/A")
            ws.cell(row, 7, "")  # Status - editable
            ws.cell(row, 8, "")  # Notes - editable
        
        # Auto-size columns
        column_widths = [12, 35, 12, 20, 15, 25, 20, 40]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_sector_sheet(self, wb: Workbook, companies: List[Company]):
        """Create sector analysis sheet."""
        ws = wb.create_sheet("Sector Analysis")
        
        # Headers
        ws['A1'] = "Sector"
        ws['B1'] = "Company Count"
        ws['C1'] = "Avg Moat Score"
        ws['D1'] = "Total Revenue"
        ws['E1'] = "HOT Count"
        
        for col in range(1, 6):
            cell = ws.cell(1, col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Group by sector
        sectors = {}
        for company in companies:
            sector = company.sector or "Unknown"
            if sector not in sectors:
                sectors[sector] = []
            sectors[sector].append(company)
        
        # Data rows
        row = 2
        for sector, sector_companies in sorted(sectors.items(), key=lambda x: len(x[1]), reverse=True):
            ws.cell(row, 1, sector)
            ws.cell(row, 2, len(sector_companies))
            ws.cell(row, 3, f"{sum(c.moat_score for c in sector_companies) / len(sector_companies):.1f}")
            ws.cell(row, 4, self._format_revenue(sum(c.revenue_gbp or 0 for c in sector_companies)))
            ws.cell(row, 5, sum(1 for c in sector_companies if c.moat_score >= 70))
            row += 1
        
        # Auto-size columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
